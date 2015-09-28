import openpyxl
from openpyxl import *

#Load your excel spreadsheet
wb = load_workbook('UpdatedList.xlsx')

#Select the worksheet, in this case the only one.
ws = wb.active

#very non-pythonic way of grouping the letters per the YaYas directory
FirstFile = ['A', 'B', 'C']
SecondFile = ['D', 'E', 'F']
ThirdFile = ['G', 'H', 'I', 'J'] 
FourthFile = ['K', 'L', 'M', 'N']
FifthFile = ['O', 'P', 'Q', 'R', 'S']
SixthFile = ['T', 'U', 'V', 'W', 'X', 'Y', 'Z']
files = [FirstFile, SecondFile, ThirdFile, FourthFile, FifthFile, SixthFile]

#clear out each output file
for file in files:
    text_file = open(str(file[0] + '-' + file[-1] +'.txt'), "w") #files are named "X-Y.txt" where X is the first letter of the range and Y is the last
    text_file.write('')
    text_file.close()

def GetLastName(num):
    return ws['A' + str(num)].value.rstrip("()0123456789) ") #need to strip some extra info from last name column

def GetFirstName(num):
    return ws['B' + str(num)].value

def GetAddress1(num):
    return ws['C' + str(num)].value

def GetAddress2(num):
    return ws['D' + str(num)].value + ", " + ws['E' + str(num)].value + " " + str(ws['F' + str(num)].value)

def GetHomeNumber(num):
    return ws['G' + str(num)].value

def GetCellNumber(num):
    return ws['H' + str(num)].value

def GetEmail(num):
    return ws['J' + str(num)].value

def GetEmail2(num):
    return ws['K' + str(num)].value.partition(",")[2].strip('")')

def GetMemberStatus(num):
    return ws['L' + str(num)].value

def GetMemberYear(num):
    return ws['M' + str(num)].value


for row in range(2, 144):
    
    currentName = GetFirstName(row) + " " + GetLastName(row)
    currentHomeNumber = GetHomeNumber(row)
    currentCell = GetCellNumber(row)
    currentAddress1 = GetAddress1(row)
    currentAddress2 = GetAddress2(row)
    currentEmail = GetEmail(row)
    currentYear = str(GetMemberYear(row))


    '''
    If the cell is empty, NoneType is returned. We need to make these NoneTypes 
    into blank strings so they can be put together with the html tags.
    '''
    if currentHomeNumber is None:
        currentPhone = currentCell
        currentHomeNumber = ""

    if currentCell is None:
        currentPhone = currentHomeNumber
        currentHomeNumber = ""

    if currentCell is not None and currentHomeNumber is not None:
        currentPhone = currentHomeNumber + " / " + currentCell

    if currentAddress1 is None:
        currentAddress1 = ""

    #very non-pythonic way of selecting our file to open based on last name; aka alphabetically sorting
    if GetLastName(row)[0] in FirstFile:
        fileToOpen = "A-C.txt"

    if GetLastName(row)[0] in SecondFile:
        fileToOpen = "D-F.txt"

    if GetLastName(row)[0] in ThirdFile:
        fileToOpen = "G-J.txt"

    if GetLastName(row)[0] in FourthFile:
        fileToOpen = "K-N.txt"

    if GetLastName(row)[0] in FifthFile:
        fileToOpen = "O-S.txt"

    if GetLastName(row)[0] in SixthFile:
        fileToOpen = "T-Z.txt"

    text_file = open(fileToOpen, "a")


    if ws['A' + str(row)].font.color.rgb == 'FFFF00FF': #If name is in pink on spreadsheet, they are a special case and need additioanl HTML
        text_file.write( '<tr>\n' +
                '    <td width="138">\n' +
                '        <p style="margin-right: 10px; margin-top: 5px; margin-bottom: 5px">\n' +
	            '            <img border="0" src="photos_members/_' + GetFirstName(row) + '_' + GetLastName(row).replace("'", "") + '.jpg' + '" width="128" height="171">\n' +
                '    </td>\n' +
	            '    <td width="298">\n'
	            '        <p style="margin-top: 0px; margin-bottom: 0; margin-left:5px" align="left">\n'
                '            <font face="Times New Roman" size="5" color="#FF8598"><img border="0" src="images/90_survivor.jpg" alt="Survivor"><br clear="right">' + currentName + '</font>\n' + 
                '            <font face="Times New Roman" size="2" color="#FF8598">&nbsp;- ' + currentYear + ' </font><br>\n' +
	            '            <font color="#FFFFFF" face="Verdana" size="2">' + currentAddress1 + ' <br>\n' +
                '                ' + currentAddress2 + '<br>\n' +
                '                ' + currentPhone + '<br>\n' +
	            '                <a href="mailto:' + currentEmail + '">' + currentEmail + '</a>\n' +
                '            </font>\n'+
                '    </td>\n' +
	            '</tr>\n')
    else: #non-pink names for the standard html
        text_file.write( '<tr>\n' +
                '    <td width="138">\n' +
                '        <p style="margin-right: 10px; margin-top: 5px; margin-bottom: 5px">\n' +
	            '            <img border="0" src="photos_members/_' + GetFirstName(row) + '_' + GetLastName(row).replace("'", "") + '.jpg' + '" width="128" height="171">\n' +
                '    </td>\n' +
	            '    <td width="298">\n'
	            '        <p style="margin-top: 0px; margin-bottom: 0; margin-left:5px" align="left">\n'
                '            <font face="Times New Roman" size="5" color="#FF8598">' + currentName + '</font>\n' + 
                '            <font face="Times New Roman" size="2" color="#FF8598">&nbsp;- ' + currentYear + ' </font><br>\n' +
	            '            <font color="#FFFFFF" face="Verdana" size="2">' + currentAddress1 + ' <br>\n' +
                '                ' + currentAddress2 + '<br>\n' +
                '                ' + currentPhone + '<br>\n' +
	            '                <a href="mailto:' + currentEmail + '">' + currentEmail + '</a>\n' +
                '            </font>\n'+
                '    </td>\n' +
	            '</tr>\n')

    text_file.close()

